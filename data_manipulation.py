from math import log2
import numpy as np
from openpyxl import Workbook, load_workbook
import re
import sbd_utils as sbd
import word_utils as word

import corpus_statistics.cs_config as cs_cfg


PARAGRAPH_TOKENIZING_RE = re.compile('\\s*\\n\\s*\\n\\s*')


def bilou_to_annotations(tags, token_indices):
    annotations = []
    annotation = None
    for tag, index_pair in zip(tags, token_indices):
        tag_mod = 'O'
        tag_type = None
        start, end = index_pair
        if tag_mod != tag:
            tag_mod, tag_type = tag.split('-')
        if tag_mod == 'O':
            if annotation:
                annotations.append(annotation)
                annotation = None
        else:
            if annotation:
                if annotation['type'] == tag_type:
                    annotation['end'] = end
                else:
                    annotations.append(annotation)
                    annotation = {'start': start, 'end': end,
                                  'type': tag_type}
            else:
                annotation = {'start': start, 'end': end, 'type': tag_type}
    if annotation and annotations and annotations[-1] != annotation:
        annotations.append(annotation)
    return annotations


def d_term_counts2list(d_term_counts):
    term_list = []
    for term, count in d_term_counts.items():
        term_list += [term]*count
    return term_list


def get_term_type(term):
    if term in cs_cfg.SS:
        return 'Small sparse'
    if term in cs_cfg.SD:
        return 'Small dense'
    if term in cs_cfg.LS:
        return 'Large sparse'
    if term in cs_cfg.LD:
        return 'Large dense'


def ndgc_score(predicted_ranking, labels_dict, k: int):
    predicted_ranking = [labels_dict[l] for l in predicted_ranking]
    gold_ranking = sorted(predicted_ranking, reverse=True)
    return dgc_score(predicted_ranking[:k]) / dgc_score(gold_ranking[:k])


def dgc_score(ranked_scores):
    return sum([s/(log2(i+1)) for i, s in enumerate(ranked_scores, 1)])


def human2cpu_label(human_label):
    return {
        "no value": 0,
        "potential value": 1,
        "certain value": 2,
        "high value": 3,
        "": np.nan
    }[human_label]


def human2cpu_label_v2(human_label):
    return {
        "no value": 0,
        "potential value": 1,
        "certain value": 5,
        "high value": 10,
        "": np.nan
    }[human_label]


def results2excel(id2score, id2text, id2label, toi, wb=None):
    ws1 = wb.create_sheet(toi)
    ordered_scores = [(sent_id, score) for sent_id, score in id2score.items()]
    ordered_scores.sort(key=lambda s: s[1], reverse=True)
    ws1['A1'] = 'Sentence'
    ws1['B1'] = 'Value'
    ws1['C1'] = 'Score'
    for i, (sent_id, score) in enumerate(ordered_scores):
        ws1['A' + str(i+2)] = id2text[sent_id]
        ws1['B' + str(i+2)] = id2label[sent_id]
        ws1['C' + str(i+2)] = score


def generate_param_pool(param_grid):

    param_pool = []
    for p, vals in param_grid.items():
        p_list = [p + '_' + str(val) for val in vals]
        if not param_pool:
            param_pool.extend([p_entry] for p_entry in p_list)
        else:
            new_param_pool = []
            for p_entry in p_list:
                for pp_entry in param_pool:
                    new_param_pool.append([p_entry] + pp_entry)
            param_pool = new_param_pool

    for param_list in param_pool:
        param_dict = {}
        for p_entry in param_list:
            k, v = p_entry.split('_')
            try:
                param_dict[k] = int(v)
            except ValueError:
                try:
                    param_dict[k] = float(v)
                except ValueError:
                    param_dict[k] = v
        yield param_dict


def sent_list2pos_list(sent_dict, para_dict, opn_dict):
    sent_list = []
    for s_id, sent in sent_dict.items():
        sent_list.append({
            '_id': s_id,
            'position': get_sent_position(sent, para_dict, opn_dict)
        })
    sorted_sent_list = sorted(sent_list, key=lambda s: s['position'])
    return [s['_id'] for s in sorted_sent_list]


def get_sent_position(sent, para_dict, opn_dict):
    opn_pos, para_pos = 0, 0
    if opn_dict:
        opn_pos = opn_dict[sent['opinion_id']]
    if para_dict:
        para_pos = para_dict[sent['paragraph_id']]
    return opn_pos, para_pos, sent['position']


def para_dict2pos_list(para_dict, opn_dict):
    para_list = []
    for p_id, para in para_dict.items():
        para_list.append({
            '_id': p_id,
            'position': get_para_position(para, opn_dict)
        })
    sorted_para_list = sorted(para_list, key=lambda p: p['position'])
    return [p['_id'] for p in sorted_para_list]


def get_para_position(para, opn_dict):
    opn_pos = 0
    if opn_dict:
        opn_pos = opn_dict[para['opinion_id']]['position']
    return opn_pos, para['position']


def organize_results_for_plotting(results_list, out_file):
    header = ('Method', 'Type', 'NDGC@10', 'NDGC@100')
    wb_out = Workbook()
    out_sheet = wb_out.get_active_sheet()
    out_sheet.append(header)
    for result_tuple in results_list:
        results_sheet = load_workbook(result_tuple[1])['NDGC']
        for row in results_sheet.iter_rows(min_row=2, values_only=True):
            out_tuple = (result_tuple[0], row[3], row[1], row[2])
            out_sheet.append(out_tuple)
    wb_out.save(out_file)


def organize_results_for_statistical_testing(results_list, out_file):
    header = ['"' + result_tuple[0] + '"' for result_tuple in results_list]
    result_dict = {}
    for result_tuple in results_list:
        results_sheet = load_workbook(result_tuple[1])['NDGC']
        for row in results_sheet.iter_rows(min_row=2, values_only=True):
            toi = row[0]
            ndgc_score = row[2]
            if toi not in result_dict:
                result_dict[toi] = []
            result_dict[toi].append(ndgc_score)
    with open(out_file, 'w') as out_f:
        out_f.write(','.join(header) + '\n')
        for toi, stats in result_dict.items():
            out_f.write(','.join(str(s) for s in stats) + ',\n')


def paragraph2features(paragraph_offsets, curr_idx, text, surr=3, call_num=0,
                       boundary_len=5):
    features = {}
    prefix = str(call_num) + ':' + 'par:'
    start, end = paragraph_offsets[curr_idx]
    para_text = text[start:end]
    ll_tokens = word.text2low_level_tokens(para_text)
    features.update({prefix+'tkn:' + tkn.lower(): 1.0 for tkn in
                    set(tkn for tkn in ll_tokens)})
    sentences = sbd.text2sentences(para_text)
    words_pos = set()
    for sentence in sentences:
        words_pos.update(sbd.pos_tag_sent(word.text2words(sentence)))
    features.update({prefix+'lemma:' +
                     word.lemmatize(lemma.lower(), pos)+'/'+pos: 1.0
                    for lemma, pos
                    in words_pos})
    features[prefix+'position'] = curr_idx / len(paragraph_offsets)
    features[prefix+'num_sent'] = float(len(sentences))
    features[prefix+'num_tkn'] = float(len(ll_tokens))
    features[prefix+'avg_sent_len'] = len(ll_tokens)/len(sentences)

    # boundary tokens
    bound_prefix = prefix + 'b:'
    for j in range((-1)*min(boundary_len, len(ll_tokens)),
                   min(boundary_len, len(ll_tokens))):
        bound_features = {bound_prefix+k: v for k, v
                          in word.token2features(ll_tokens[j], j).items()}
        features.update(bound_features)
    return features


def text2paragraphs(text, offsets=False):
    separators = PARAGRAPH_TOKENIZING_RE.finditer(text)
    end = 0
    paragraph_offsets = []
    for sep in separators:
        start = sep.start()
        if start-end > 0:
            paragraph_offsets.append((end, start))
        end = sep.end()
    start = len(text)
    if start-end > 0:
        paragraph_offsets.append((end, start))
    if offsets:
        return paragraph_offsets
    return [text[start:end] for start, end in paragraph_offsets]


def normalize_cached_scores(cache_dict):
    max_val = None
    min_val = None
    norm_dict = {}
    for toi, data in cache_dict.items():
        for item_id, score in data.items():
            if max_val is None:
                max_val = score
            if min_val is None:
                min_val = score
            if max_val < score:
                max_val = score
            if min_val > score:
                min_val = score

    for toi, data in cache_dict.items():
        norm_dict[toi] = {}
        for item_id, score in data.items():
            norm_dict[toi][item_id] = (score - min_val) / (max_val - min_val)

    return norm_dict


def normalize_toi_scores(snt_id2score):
    max_val = None
    min_val = None
    norm_dict = {}
    for item_id, score in snt_id2score.items():
        if max_val is None or max_val < score:
            max_val = score
        if min_val is None or min_val > score:
            min_val = score
    for item_id, score in snt_id2score.items():
        norm_dict[item_id] = (score - min_val) / (max_val - min_val)

    return norm_dict
